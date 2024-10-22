import streamlit as st
import pandas as pd
import os
from datetime import timedelta
import altair as alt
import openpyxl
from openpyxl.styles import NamedStyle
from datetime import datetime
import yfinance as yf
import subprocess
from openpyxl.utils import get_column_letter # To run git commands

# Define the directory where the workbooks are stored (this is in the same repo)
WORKBOOK_DIR = "NAV"  # Folder where the Excel workbooks are stored

# Function to list available Excel files in the specified directory
def list_workbooks(directory):
    try:
        # List only .xlsx files in the directory
        files = [f for f in os.listdir(directory) if f.endswith('.xlsx')]
        return files
    except FileNotFoundError:
        st.error("Directory not found. Please ensure the specified directory exists.")
        return []

# Function to load NAV data from the selected workbook
def load_nav_data(file_path):
    try:
        # Load full sheet data without limiting columns and headers
        data = pd.read_excel(file_path, sheet_name=0, header=None)
        data.columns = ['Date', 'Header', 'Stock1', 'Stock2', 'Stock3', 'Stock4', 'Stock5', 'Basket Value', 'Returns', 'NAV']
        
        # Ensure 'Date' column is datetime; coerce errors to handle non-date values
        data['Date'] = pd.to_datetime(data['Date'], errors='coerce')
        

        # Check if required columns are present
        if 'NAV' not in data.columns or 'Date' not in data.columns:
            st.error("NAV or Date column not found in the selected workbook.")
            return pd.DataFrame()

        return data
    except Exception as e:
        st.error(f"Error reading Excel file: {e}")
        return pd.DataFrame()

# Function to filter data based on the selected date range
def filter_data_by_date(data, date_range):
    if date_range == "1 Day":
        return data.tail(1)
    elif date_range == "5 Days":
        return data.tail(5)
    elif date_range == "1 Month":
        one_month_ago = data['Date'].max() - timedelta(days=30)
        return data[data['Date'] >= one_month_ago]
    elif date_range == "6 Months":
        six_months_ago = data['Date'].max() - timedelta(days=180)
        return data[data['Date'] >= six_months_ago]
    elif date_range == "1 Year":
        one_year_ago = data['Date'].max() - timedelta(days=365)
        return data[data['Date'] >= one_year_ago]
    else:  # Max
        return data

# Function to process Excel data and identify stock name changes dynamically
def process_excel_data(data):
    stock_blocks = []
    current_block = None

    # Iterate through the rows of the DataFrame to detect stock changes
    for idx, row in data.iterrows():
        if isinstance(row['Header'], str) and row['Header'] == 'Stocks':  # Detect when stock names change
            if current_block:
                current_block['end_idx'] = idx - 1  # End the current block before the next 'Stocks' row
                stock_blocks.append(current_block)  # Save the completed block

            # Create a new block
            stock_names = row[['Stock1', 'Stock2', 'Stock3', 'Stock4', 'Stock5']].tolist()  # Get stock names from columns C to G
            current_block = {'stock_names': stock_names, 'start_idx': idx + 2, 'end_idx': None, 'dates': []}  # Include dates
           

    if current_block:
        current_block['end_idx'] = len(data) - 1  # Handle the last block until the end of the dataset
        stock_blocks.append(current_block)

    # Add dates to each stock block
    for block in stock_blocks:
        block['dates'] = data.iloc[block['start_idx']:block['end_idx'] + 1].dropna(subset=['Date'])['Date'].tolist()
        

    return stock_blocks

# Function to insert stock names for the relevant block above the selected time period's data
def insert_stock_names_above_data(stock_blocks, filtered_data):
    final_data = pd.DataFrame()

    filtered_dates = filtered_data['Date'].tolist()

    for block in stock_blocks:
        # Check if any dates from this block overlap with filtered dates
        overlap_dates = [date for date in block['dates'] if date in filtered_dates]

        # If there are overlapping dates, insert the stock names above the first relevant date
        if overlap_dates:
            # Insert the stock names once, before the first matching date in the block
            stock_names_row = pd.DataFrame([[None] * len(filtered_data.columns)], columns=filtered_data.columns)
            for i, stock_name in enumerate(block['stock_names']):
                stock_names_row[f'Stock{i + 1}'] = stock_name

            # Add the stock names row to the final data
            final_data = pd.concat([final_data, stock_names_row], ignore_index=True)

            # Add the block's data that overlaps with the filtered data
            block_data = filtered_data[filtered_data['Date'].isin(overlap_dates)]
            final_data = pd.concat([final_data, block_data], ignore_index=True)

    return final_data

def highlight_stock_names(df, highlighted_rows):
    def highlight_row(row):
        # If the index of the row is in highlighted_rows, apply background color
        return ['background-color: yellow' if row.name in highlighted_rows else '' for _ in row]

    return df.style.apply(highlight_row, axis=1)

# Function to recalculate NAV starting from 100
def recalculate_nav(filtered_data):
    initial_nav = filtered_data['NAV'].iloc[0]
    filtered_data['Rebased NAV'] = (filtered_data['NAV'] / initial_nav) * 100
    return filtered_data

# Function to modify all Excel files in the directory and push them to GitHub
def modify_all_workbooks_and_push_to_github():
    workbooks = list_workbooks(WORKBOOK_DIR)
    if not workbooks:
        st.error("No workbooks found to modify.")
        return

    modified_files = []

    for filename in workbooks:
        try:
            modify_workbook(filename)
            modified_files.append(filename)
        except Exception as e:
            st.error(f"Error modifying {filename}: {e}")

    # Push all modified files to GitHub
    if modified_files:
        git_add_commit_push(modified_files)

# Function to modify a single Excel workbook
def modify_workbook(filename):
    file_path = os.path.join(WORKBOOK_DIR, filename)
    try:
        workbook = openpyxl.load_workbook(file_path)

        # Create a style for date formatting
        date_style = NamedStyle(name="datetime", number_format='yyyy-mm-dd')
        
        if "datetime" not in workbook.named_styles:
            workbook.add_named_style(date_style)
    
        for sheet_name in workbook.sheetnames:
            ws = workbook[sheet_name]
            print(f"Modifying sheet: {sheet_name}")

            # Step 1: Find the actual last row with data in the worksheet
            last_row = ws.max_row
            while last_row > 1 and ws.cell(row=last_row, column=1).value in (None, ""):
                last_row -= 1

            # Step 2: Determine the next date to add
            last_date = None
            for row in range(last_row, 1, -1):
                cell_value = ws.cell(row=row, column=1).value
                if isinstance(cell_value, datetime):
                    last_date = cell_value
                    break
                elif isinstance(cell_value, str):
                    try:
                        last_date = parser.parse(cell_value)
                        break
                    except ValueError:
                        continue  # Skip rows that cannot be parsed as a date

            if last_date is None:
                # If no valid date is found, set a fallback date
                last_date = datetime.now() - timedelta(days=1)

            next_date = last_date + timedelta(days=1)

            # Step 3: Identify the last non-zero NAV in column J (NAV)
            nav_column_index = 10
            last_non_zero_nav = None

            for row in range(last_row, 1, -1):
                nav_value = ws.cell(row=row, column=nav_column_index).value
                if isinstance(nav_value, (int, float)) and nav_value != 0:
                    last_non_zero_nav = nav_value
                    break

            if last_non_zero_nav is None:
                last_non_zero_nav = 100

            # Step 4: Identify existing stock symbols and quantities in columns C to G
            stocks_row = None
            quantities_row = None

            for row in range(1, ws.max_row + 1):
                cell_value = ws.cell(row=row, column=2).value
                if cell_value == "Stocks":
                    stocks_row = row
                elif cell_value == "Quantities":
                    quantities_row = row

            if not stocks_row or not quantities_row:
                print(f"Could not find 'Stocks' or 'Quantities' headers in sheet {sheet_name}. Skipping sheet.")
                continue

            stocks = {}
            quantities = []

            for col in range(3, 8):
                stock_symbol = ws.cell(row=stocks_row, column=col).value
                quantity = ws.cell(row=quantities_row, column=col).value
                if stock_symbol and isinstance(stock_symbol, str):
                    stocks[stock_symbol] = stock_symbol
                    quantities.append(quantity)

            # Step 5: Fetch historical stock data
            today_date = datetime.now().strftime('%Y-%m-%d')
            next_date_str = next_date.strftime('%Y-%m-%d')

            all_prices = {}
            for stock_symbol in stocks.keys():
                ticker = yf.Ticker(stock_symbol)

                try:
                    hist = ticker.history(start=next_date_str, end=today_date, interval="1d", auto_adjust=False)
                    if hist.empty:
                        print(f"No data found for {stock_symbol}. Skipping.")
                        continue

                    closing_prices = hist['Close'].tolist()
                    closing_dates = hist.index.strftime('%Y-%m-%d').tolist()
                    all_prices[stock_symbol] = (closing_dates, closing_prices)

                except Exception as e:
                    print(f"Error fetching data for {stock_symbol}: {e}")
                    continue

            # Step 6: Insert the fetched data and perform calculations
            basket_values = []
            returns = []
            nav_values = [last_non_zero_nav]

            for i in range(len(closing_dates)):
                # Convert the current date to datetime.date for comparison
                current_date = datetime.strptime(closing_dates[i], '%Y-%m-%d').date()

                # Check if the date already exists in the worksheet
                if any(ws.cell(row=r, column=1).value == current_date for r in range(2, last_row + 1)):
                    print(f"Date {current_date} already exists. Skipping.")
                    continue

                current_row = last_row + 1  # Add data to the immediate next row after the last data row
                last_row += 1

                # Insert date
                date_value = datetime.strptime(closing_dates[i], '%Y-%m-%d')
                date_cell = ws.cell(row=current_row, column=1, value=date_value)
                date_cell.number_format = 'yyyy-mm-dd'  # Apply the date style to the cell

                # Calculate basket value for the current date
                basket_value = 0
                for j, stock_symbol in enumerate(stocks.keys()):
                    price = all_prices[stock_symbol][1][i] if i < len(all_prices[stock_symbol][1]) else 0
                    quantity = quantities[j]
                    basket_value += price * quantity
                    ws.cell(row=current_row, column=3 + j, value=price)  # Insert price starting from column C

                # Insert basket value in column H
                ws.cell(row=current_row, column=8, value=basket_value)
                basket_values.append(basket_value)

                # Calculate returns and insert in column I
                ret = (basket_value - basket_values[i - 1]) / basket_values[i - 1] if i > 0 and basket_values[i - 1] != 0 else 0
                returns.append(ret)
                ws.cell(row=current_row, column=9, value=ret)

                # Calculate NAV and insert in column J
                nav = nav_values[-1] * (1 + ret)
                nav_values.append(nav)
                ws.cell(row=current_row, column=10, value=nav)

        workbook.save(file_path)

    except Exception as e:
        print(f"Error modifying {filename}: {e}")

# Function to execute git commands to add, commit, and push changes
def git_add_commit_push(modified_files):
    try:
        # Git add each modified file
        for filename in modified_files:
            subprocess.run(["git", "add", f"{WORKBOOK_DIR}/{filename}"], check=True)

        # Check if there are changes to commit
        status_result = subprocess.run(["git", "status", "--porcelain"], capture_output=True, text=True, check=True)
        
        # If there are no changes, return without committing
        if not status_result.stdout.strip():
            st.warning("No changes to commit.")
            return

        # Git commit with a single message for all files
        commit_message = f"Updated {', '.join(modified_files)} with new data"
        subprocess.run(["git", "commit", "-m", commit_message], check=True)

        # Git push to the remote repository
        subprocess.run(["git", "push"], check=True)

    except subprocess.CalledProcessError as e:
        print(f"Error during git operation: {e}")

def clean_chart_data(filtered_data, chart_column):
    # Convert the chart_column to numeric, replacing non-numeric values with NaN
    filtered_data[chart_column] = pd.to_numeric(filtered_data[chart_column], errors='coerce')
    
    # Drop rows where the chart_column is NaN
    clean_data = filtered_data.dropna(subset=[chart_column])
    return clean_data

def format_table_data(data):
    # Round numeric columns to 2 decimal places
    for col in ['Stock1', 'Stock2', 'Stock3', 'Stock4', 'Stock5', 'Basket Value', 'Returns', 'NAV']:
        data[col] = pd.to_numeric(data[col], errors='coerce').round(3).fillna(data[col])
    # Format date to exclude time
    data['Date'] = data['Date'].dt.strftime('%Y-%m-%d')
    
    return data


def main():
    st.title("NAV Data Dashboard")
    modify_all_workbooks_and_push_to_github()

    # List available workbooks in the directory
    workbooks = list_workbooks(WORKBOOK_DIR)

    if not workbooks:
        st.error("No Excel workbooks found in the specified directory.")
        return

    # Display the data for a specific workbook (example: the first one)
    selected_workbook = st.selectbox("Select a workbook", workbooks)
    
    file_path = os.path.join(WORKBOOK_DIR, selected_workbook)

    nav_data = load_nav_data(file_path)

    if not nav_data.empty:
        # Process the Excel data and detect stock name changes, including block dates
        stock_blocks = process_excel_data(nav_data)

        if not stock_blocks:
            st.error("No valid stock data found in the workbook.")
            return

        # Allow the user to select a date range
        date_ranges = ["1 Day", "5 Days", "1 Month", "6 Months", "1 Year", "Max"]
        selected_range = st.selectbox("Select Date Range", date_ranges)

        # Filter the nav_data by the selected date range
        filtered_data = filter_data_by_date(nav_data, selected_range)
        if selected_range not in ["1 Day", "Max"]:
            filtered_data = recalculate_nav(filtered_data)
            chart_column = 'Rebased NAV'
        else:
            chart_column = 'NAV'

        clean_filtered_data = clean_chart_data(filtered_data, chart_column)


        line_chart = alt.Chart(clean_filtered_data).mark_line().encode(
            x='Date:T',
            y=alt.Y(f'{chart_column}:Q', scale=alt.Scale(domain=[80, clean_filtered_data[chart_column].max()])),
            tooltip=['Date:T', f'{chart_column}:Q']
        ).properties(
            width=700,
            height=400
        )
        st.write(f"### Displaying data from {selected_workbook}")
        st.altair_chart(line_chart, use_container_width=True)

        # Insert stock names above the relevant block data
        final_data = insert_stock_names_above_data(stock_blocks, filtered_data)
        final_data = format_table_data(final_data)

        
        # Display the combined filtered data with highlighted stock names
        st.write("### Stock Data Table")
        st.dataframe(final_data)


    else:
        st.error("Failed to load data. Please check the workbook format.")

if __name__ == "__main__":
    main()
